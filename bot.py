from telegram import Update, BotCommand
from telegram.ext import Application, MessageHandler, CommandHandler, filters, ContextTypes
from openpyxl import load_workbook
import logging
import os
import requests

logging.basicConfig(level=logging.INFO)
token = os.getenv("TELEGRAM_BOT_TOKEN")
onedrive_url = os.getenv("ONEDRIVE_EXCEL_URL")
RENDER_URL = os.getenv("RENDER_EXTERNAL_URL")
PORT = int(os.getenv("PORT", "10000"))

def calcular_total(ws):
    montos = []
    fila = 3
    while ws[f'B{fila}'].value is not None:
        val = ws[f'B{fila}'].value
        if isinstance(val, (int, float)):
            montos.append(val)
        fila += 1
    f3 = ws['F3'].value
    f3 = f3 if isinstance(f3, (int, float)) else 0
    return montos, sum(montos) + f3

async def start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.message.reply_text(
        "Comandos disponibles:\n"
        "/total - Ver total de gastos\n"
        "/gastos - Ver todos los gastos\n"
        "\nO simplemente envía un número para agregar un gasto"
    )

async def ver_total(update: Update, context: ContextTypes.DEFAULT_TYPE):
    try:
        response = requests.get(onedrive_url)
        with open('Tarjeta de Credito.xlsx', 'wb') as f:
            f.write(response.content)

        wb = load_workbook('Tarjeta de Credito.xlsx')
        ws = wb['21 de agosto']
        _, total = calcular_total(ws)

        await update.message.reply_text(f"💰 Total de gastos: ${total}")
    except Exception as e:
        await update.message.reply_text(f"❌ Error: {str(e)}")

async def ver_gastos(update: Update, context: ContextTypes.DEFAULT_TYPE):
    try:
        response = requests.get(onedrive_url)
        with open('Tarjeta de Credito.xlsx', 'wb') as f:
            f.write(response.content)

        wb = load_workbook('Tarjeta de Credito.xlsx')
        ws = wb['21 de agosto']
        montos, total = calcular_total(ws)

        lista = '\n'.join([f"• ${g}" for g in montos])
        await update.message.reply_text(f"📊 Gastos:\n{lista}\n\n💰 Total: ${total}")
    except Exception as e:
        await update.message.reply_text(f"❌ Error: {str(e)}")

async def recibir_monto(update: Update, context: ContextTypes.DEFAULT_TYPE):
    try:
        monto = float(update.message.text)

        response = requests.get(onedrive_url)
        with open('Tarjeta de Credito.xlsx', 'wb') as f:
            f.write(response.content)

        wb = load_workbook('Tarjeta de Credito.xlsx')
        ws = wb['21 de agosto']

        fila = 3
        while ws[f'B{fila}'].value is not None:
            fila += 1

        ws[f'B{fila}'].value = monto
        wb.save('Tarjeta de Credito.xlsx')

        with open('Tarjeta de Credito.xlsx', 'rb') as f:
            requests.put(onedrive_url, data=f)

        _, total = calcular_total(ws)

        await update.message.reply_text(f"✅ Agregado: ${monto}\n\n💰 Total: ${total}")
    except ValueError:
        await update.message.reply_text("❌ Envía solo un número")
    except Exception as e:
        await update.message.reply_text(f"❌ Error: {str(e)}")

async def post_init(application: Application):
    await application.bot.set_my_commands([
        BotCommand("start", "Ver ayuda"),
        BotCommand("total", "Ver total de gastos"),
        BotCommand("gastos", "Ver todos los gastos"),
    ])

if __name__ == '__main__':
    app = Application.builder().token(token).post_init(post_init).build()
    app.add_handler(CommandHandler("start", start))
    app.add_handler(CommandHandler("total", ver_total))
    app.add_handler(CommandHandler("gastos", ver_gastos))
    app.add_handler(MessageHandler(filters.TEXT & ~filters.COMMAND, recibir_monto))

    app.run_webhook(
        listen="0.0.0.0",
        port=PORT,
        url_path=token,
        webhook_url=f"{RENDER_URL}/{token}"
    )